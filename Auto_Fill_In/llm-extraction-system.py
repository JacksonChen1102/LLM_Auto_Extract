import os
import time
import logging
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional, Any
import json
import sys
import re
from datetime import datetime
import io
import tempfile
import fitz  # PyMuPDF

# 设置日志
# 解决Windows命令行中文编码问题
if sys.platform == 'win32':
    # 仅将日志输出到文件，避免控制台编码问题
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler("extraction_log.log", encoding='utf-8')
        ]
    )
    # 添加一个简单的英文控制台处理器
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - Processing... (see log file for details)'))
    logging.getLogger('').addHandler(console)
else:
    # 非Windows平台使用标准配置
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler("extraction_log.log", encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
logger = logging.getLogger(__name__)


class DataFormatter:
    """数据格式化工具：处理日期格式和Yes/No转换"""

    @staticmethod
    def format_date(date_str):
        """
        将各种日期格式转换为yyyy-mm-dd格式

        Args:
            date_str: 原始日期字符串

        Returns:
            格式化后的日期字符串，如果无法解析则返回原值
        """
        if not date_str or pd.isna(date_str):
            return ""

        date_str = str(date_str).strip()

        # 尝试不同的日期格式匹配和转换
        date_formats = [
            # 标准格式
            "%Y-%m-%d",
            # 常见替代格式
            "%d %B %Y", "%d %b %Y", "%B %d, %Y", "%b %d, %Y",
            "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d",
            # 日期+月份名称格式
            "%d %B, %Y", "%d %b, %Y", "%d.%B.%Y", "%d.%b.%Y",
            # 仅包含日期和月份名称的格式
            "%d %B", "%d %b", "%B %d", "%b %d",
            # 特殊格式
            "%dth of %B, %Y", "%dth of %B %Y", "%dst of %B %Y", "%dnd of %B %Y", "%drd of %B %Y",
            "%dth of %b, %Y", "%dth of %b %Y", "%dst of %b %Y", "%dnd of %b %Y", "%drd of %b %Y",
            # 其他格式
            "%d.%m.%y", "%d/%m/%y", "%m/%d/%y", "%y/%m/%d"
        ]

        # 尝试使用正则表达式处理特殊格式
        # 处理"9th of June, 2025"这样的格式
        special_format = re.match(r'(\d+)(?:st|nd|rd|th) of (\w+),? (\d{4})', date_str)
        if special_format:
            day, month, year = special_format.groups()
            date_str = f"{day} {month} {year}"

        # 处理"26.Jun.2025"这样的格式
        dot_format = re.match(r'(\d+)\.(\w+)\.(\d{4})', date_str)
        if dot_format:
            day, month, year = dot_format.groups()
            date_str = f"{day} {month} {year}"

        # 处理"15 July, 2025"这样的格式
        comma_format = re.match(r'(\d+) (\w+), (\d{4})', date_str)
        if comma_format:
            day, month, year = comma_format.groups()
            date_str = f"{day} {month} {year}"

        # 尝试解析为日期对象
        for fmt in date_formats:
            try:
                # 尝试将字符串解析为日期对象
                date_obj = datetime.strptime(date_str, fmt)
                # 如果没有年份信息，假设为当前年
                if '%Y' not in fmt and '%y' not in fmt:
                    date_obj = date_obj.replace(year=datetime.now().year)
                # 返回标准格式
                return date_obj.strftime('%Y-%m-%d')
            except ValueError:
                continue

        # 使用正则表达式尝试匹配常见的日期模式
        # 匹配"Tuesday 30 September 2025"这样的格式
        weekday_pattern = re.search(r'(\w+) (\d+) (\w+) (\d{4})', date_str)
        if weekday_pattern:
            try:
                _, day, month, year = weekday_pattern.groups()
                temp_str = f"{day} {month} {year}"
                for fmt in ["%d %B %Y", "%d %b %Y"]:
                    try:
                        date_obj = datetime.strptime(temp_str, fmt)
                        return date_obj.strftime('%Y-%m-%d')
                    except ValueError:
                        continue
            except Exception:
                pass

        # 如果所有尝试都失败，返回原字符串
        logger.warning(f"无法解析日期格式: {date_str}")
        return date_str

    @staticmethod
    def normalize_value(value):
        """
        将值标准化：Yes -> 1, No -> ""

        Args:
            value: 原始值

        Returns:
            标准化后的值
        """
        if pd.isna(value) or value == "":
            return ""

        value_str = str(value).strip().lower()

        # 处理Yes/No情况
        if value_str in ['yes', 'y', 'true', '1']:
            return "1"
        elif value_str in ['no', 'n', 'false', '0']:
            return ""
        elif value_str and value_str != 'nan':  # 有其他非空值
            # 检查是否是数字（如1904）
            if re.match(r'^\d+$', value_str):
                return value_str  # 保持数字不变
            else:
                return "1"  # 非数字的其他值也转为1
        else:
            return ""

    @staticmethod
    def process_results(results: Dict[str, Any]) -> Dict[str, Any]:
        """
        处理LLM提取的结果：格式化日期，标准化值

        Args:
            results: 原始结果字典

        Returns:
            处理后的结果字典
        """
        processed_results = {}

        # 定义类别字段列表（仅这些字段需要转换为1或空白）
        category_fields = ["Master Student", "Doctoral Student", "PostDoc", "Research Assistant",
                           "Competition", "Summer School", "Conference", "Workshop",
                           "Physical_Geo", "Human_Geo", "Urban", "GIS", "RS", "GNSS"]

        # 定义日期字段
        date_fields = ["Deadline"]

        # 定义数字字段
        number_fields = ["Number_Places"]

        # 定义翻译字段
        translation_fields = ["University_CN", "Country_CN"]

        # 定义文本字段 - 通过排除法
        text_fields = [k for k in results.keys() if k not in category_fields and
                       k not in date_fields and k not in number_fields and
                       k not in translation_fields and k not in ["Notes", "Source", "Verifier", "Error"]]

        for key, value in results.items():
            # 检查键是否为日期相关字段
            if key in date_fields:
                processed_results[key] = DataFormatter.format_date(value)
            # 检查键是否为类别字段
            elif key in category_fields:
                processed_results[key] = DataFormatter.normalize_value(value)
            # 检查键是否为数量字段（保持原值，仅确保为数字）
            elif key in number_fields:
                # 如果是数字字符串，保持原样
                if value and str(value).strip().isdigit():
                    processed_results[key] = str(value).strip()
                # 处理英文数字转换为阿拉伯数字
                elif value and isinstance(value, str):
                    # 英文数字到阿拉伯数字的映射
                    english_to_number = {
                        "zero": "0", "one": "1", "two": "2", "three": "3", "four": "4",
                        "five": "5", "six": "6", "seven": "7", "eight": "8", "nine": "9",
                        "ten": "10", "eleven": "11", "twelve": "12", "thirteen": "13",
                        "fourteen": "14", "fifteen": "15", "sixteen": "16",
                        "seventeen": "17", "eighteen": "18", "nineteen": "19",
                        "twenty": "20", "thirty": "30", "forty": "40", "fifty": "50",
                        "sixty": "60", "seventy": "70", "eighty": "80", "ninety": "90"
                    }

                    value_lower = str(value).strip().lower()

                    # 直接映射
                    if value_lower in english_to_number:
                        processed_results[key] = english_to_number[value_lower]
                    # 如果不是数字且无法映射，保持原值并记录警告
                    else:
                        logger.warning(f"数量字段 '{key}' 的值 '{value}' 无法转换为数字")
                        processed_results[key] = str(value).strip()
                else:
                    processed_results[key] = ""
            # 文本字段和翻译字段保持原值，不做转换
            elif key in text_fields or key in translation_fields:
                # 确保文本值不会被误处理为1
                if value and str(value).strip().lower() in ['yes', 'y', 'true']:
                    logger.warning(f"文本字段 '{key}' 的值为 '{value}'，这可能是错误的。保持原值而不转换为1。")
                processed_results[key] = str(value).strip() if value and not pd.isna(value) else ""
            # 其他字段按默认处理
            else:
                processed_results[key] = str(value).strip() if value and not pd.isna(value) else ""

        return processed_results


class ExcelProcessor:
    """Excel处理模块：负责读取、筛选行、写回结果"""

    def __init__(self, excel_path: str, sheet_name: str = "Unfilled"):
        """
        初始化Excel处理器

        Args:
            excel_path: Excel文件路径
            sheet_name: 工作表名称，默认为"Unfilled"
        """
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.df = None
        self.workbook = None

    def load_data(self) -> pd.DataFrame:
        """
        加载Excel数据

        Returns:
            包含Excel数据的DataFrame
        """
        try:
            self.df = pd.read_excel(self.excel_path, sheet_name=self.sheet_name)
            # 加载workbook以便后续写入
            self.workbook = load_workbook(self.excel_path)
            logger.info(f"成功加载Excel文件: {self.excel_path}, 表格: {self.sheet_name}")
            return self.df
        except Exception as e:
            logger.error(f"加载Excel文件失败: {str(e)}")
            raise

    def filter_rows(self) -> pd.DataFrame:
        """
        筛选需要处理的行

        Returns:
            需要处理的行组成的DataFrame
        """
        if self.df is None:
            self.load_data()

        # 确保必要的列存在
        required_columns = ['Notes', 'Verifier', 'Error']
        for col in required_columns:
            if col not in self.df.columns:
                logger.warning(f"表格中缺少列 '{col}'，将创建该列")
                self.df[col] = None

        # 过滤需要处理的行
        # Notes 列为空 或 包含 URL
        # Verifier 列为空
        # Error 列为空
        filtered_df = self.df[
            ((self.df['Notes'].isna()) |
             (self.df['Notes'] == '') |
             (self.df['Notes'].apply(lambda x: self._is_url(x) if isinstance(x, str) else False))) &
            (self.df['Verifier'].isna() | (self.df['Verifier'] == '')) &
            (self.df['Error'].isna() | (self.df['Error'] == ''))
        ]

        logger.info(f"总行数: {len(self.df)}, 需处理的行数: {len(filtered_df)}")
        return filtered_df

    def _is_url(self, text: str) -> bool:
        """
        判断文本是否为URL

        Args:
            text: 要判断的文本

        Returns:
            是否为URL
        """
        # URL正则表达式模式
        url_pattern = re.compile(
            r'^(https?:\/\/)?(www\.)?[-a-zA-Z0-9@:%._\+~#=]{1,256}\.'
            r'[a-zA-Z0-9()]{1,6}\b([-a-zA-Z0-9()@:%_\+.~#?&//=]*)'
        )
        return bool(url_pattern.match(text.strip()))

    def write_results(self, row_index: int, results: Dict[str, Any], has_error: bool = False,
                      error_msg: str = "") -> None:
        """
        将结果写回Excel

        Args:
            row_index: 行索引
            results: 提取的结构化结果
            has_error: 是否有错误
            error_msg: 错误信息
        """
        # 获取实际的Excel行索引（因为DataFrame索引可能与Excel行号不同）
        excel_row = row_index + 2  # 加2是因为Excel有表头，且索引从1开始

        try:
            # 获取工作表
            ws = self.workbook[self.sheet_name]

            # 如果有错误，写入Error列
            if has_error:
                error_col = self._get_column_index("Error")
                ws.cell(row=excel_row, column=error_col).value = error_msg
                logger.warning(f"行 {row_index} 处理失败: {error_msg}")
            else:
                # 处理结果数据
                processed_results = DataFormatter.process_results(results)
                # 写入所有提取的结果
                for col_name, value in processed_results.items():
                    col_index = self._get_column_index(col_name)
                    if col_index:
                        ws.cell(row=excel_row, column=col_index).value = value

                # 在Verifier列填入"LLM"
                verifier_col = self._get_column_index("Verifier")
                if verifier_col:
                    ws.cell(row=excel_row, column=verifier_col).value = "LLM"
                    logger.info(f"在行 {row_index} 的Verifier列填入'LLM'")

                logger.info(f"成功写入行 {row_index} 的结果")

            # 保存工作簿
            self.workbook.save(self.excel_path)

        except Exception as e:
            logger.error(f"写入结果到Excel失败, 行 {row_index}: {str(e)}")

    def _get_column_index(self, column_name: str) -> int:
        """
        获取列名对应的列索引

        Args:
            column_name: 列名

        Returns:
            列索引（从1开始）
        """
        columns = list(self.df.columns)
        try:
            # 列索引从1开始
            return columns.index(column_name) + 1
        except ValueError:
            logger.error(f"列名 '{column_name}' 不存在")
            return 0


class WebPageExtractor:
    """网页提取模块：处理普通网页和PDF链接"""

    def __init__(self, timeout: int = 30):
        """
        初始化网页提取器

        Args:
            timeout: 请求超时时间（秒）
        """
        self.timeout = timeout
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

    def extract_content(self, url: str) -> str:
        """
        提取网页或PDF内容的主方法

        Args:
            url: 网页URL

        Returns:
            提取的内容文本
        """
        url = url.strip()

        # 判断URL类型
        if url.lower().endswith('.pdf') or '/pdf/' in url.lower():
            logger.info(f"检测到PDF链接: {url}")
            return self.extract_pdf_from_url(url)
        else:
            logger.info(f"检测到普通网页: {url}")
            return self.extract_html_page(url)

    def extract_pdf_from_url(self, url: str) -> str:
        """
        从URL下载PDF并提取文本

        Args:
            url: PDF文件的URL

        Returns:
            提取的文本内容
        """
        try:
            # 下载PDF文件
            logger.info(f"正在下载PDF: {url}")
            response = requests.get(url, headers=self.headers, timeout=self.timeout)
            response.raise_for_status()

            # 检查内容类型
            content_type = response.headers.get('Content-Type', '').lower()
            if 'application/pdf' not in content_type and not url.lower().endswith('.pdf'):
                logger.warning(f"URL返回的不是PDF内容 (Content-Type: {content_type})")
                # 如果不是PDF，尝试作为HTML处理
                return self.extract_html_content(response.text)

            # 提取PDF内容
            return self.extract_pdf_content(response.content)

        except Exception as e:
            logger.error(f"下载或处理PDF失败: {str(e)}")
            return f"[提取失败: {str(e)}]"

    def extract_pdf_content(self, pdf_data: bytes) -> str:
        """
        从PDF二进制数据中提取文本，使用PyMuPDF

        Args:
            pdf_data: PDF二进制数据

        Returns:
            提取的文本内容
        """
        temp_file = None
        temp_file_path = None
        try:
            # 创建临时文件
            temp_file = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
            temp_file_path = temp_file.name
            temp_file.write(pdf_data)
            temp_file.close()

            logger.info(f"PDF已保存到临时文件: {temp_file_path}")

            # 使用PyMuPDF提取文本
            text = ""
            with fitz.open(temp_file_path) as doc:
                logger.info(f"PDF包含 {len(doc)} 页")
                for page_num, page in enumerate(doc):
                    page_text = page.get_text()
                    text += page_text + "\n\n"
                    logger.debug(f"已提取第 {page_num + 1} 页文本，长度: {len(page_text)}")

            # 清理文本
            lines = [line.strip() for line in text.splitlines() if line.strip()]
            clean_text = '\n'.join(lines)

            logger.info(f"成功提取PDF内容，长度: {len(clean_text)}")
            return clean_text

        except Exception as e:
            logger.error(f"使用PyMuPDF提取PDF内容失败: {str(e)}")
            return f"[PDF解析失败: {str(e)}]"

        finally:
            # 确保临时文件被删除
            if temp_file_path is not None:
                try:
                    os.unlink(temp_file_path)
                    logger.debug(f"临时PDF文件已删除: {temp_file_path}")
                except Exception as e:
                    logger.warning(f"删除临时文件失败: {str(e)}")

    def extract_html_page(self, url: str) -> str:
        """
        提取普通HTML网页内容

        Args:
            url: 网页URL

        Returns:
            提取的文本内容
        """
        try:
            response = requests.get(url, headers=self.headers, timeout=self.timeout)
            response.raise_for_status()

            # 检查响应类型是否为PDF
            content_type = response.headers.get('Content-Type', '').lower()
            if 'application/pdf' in content_type:
                logger.info(f"URL返回了PDF内容: {url}")
                return self.extract_pdf_content(response.content)

            # 处理HTML内容
            return self.extract_html_content(response.text)

        except Exception as e:
            logger.error(f"提取HTML内容失败: {str(e)}")
            return f"[提取失败: {str(e)}]"

    def extract_html_content(self, html_text: str) -> str:
        """
        从HTML文本中提取内容

        Args:
            html_text: HTML文本

        Returns:
            提取的文本内容
        """
        try:
            soup = BeautifulSoup(html_text, 'html.parser')

            # 移除不需要的元素
            for element in soup(['script', 'style', 'header', 'footer', 'nav']):
                element.decompose()

            # 获取文本内容
            text = soup.get_text(separator='\n')

            # 清理文本
            lines = [line.strip() for line in text.splitlines() if line.strip()]
            clean_text = '\n'.join(lines)

            logger.info(f"成功提取HTML内容，长度: {len(clean_text)}")
            return clean_text

        except Exception as e:
            logger.error(f"处理HTML内容失败: {str(e)}")
            return f"[HTML处理失败: {str(e)}]"


class LlmProcessor:
    """LLM交互模块：将正文文本输入给本地模型，获取结构化输出"""

    def __init__(self, model_name: str = "llama3", api_base: str = "http://localhost:11434"):
        """
        初始化LLM处理器

        Args:
            model_name: Ollama模型名称
            api_base: Ollama API地址
        """
        self.model_name = model_name
        self.api_base = api_base
        self.api_url = f"{api_base}/api/generate"

    def extract_structured_info(self, text: str, fields: List[str]) -> Dict[str, str]:
        """
        使用LLM从文本中提取结构化信息

        Args:
            text: 网页正文文本
            fields: 需要提取的字段列表

        Returns:
            提取的结构化信息字典
        """
        # 构建提示词
        prompt = self._build_prompt(text, fields)

        try:
            logger.info(f"正在使用LLM提取信息，模型: {self.model_name}")

            # 调用Ollama API
            payload = {
                "model": self.model_name,
                "prompt": prompt,
                "stream": False,
                "options": {
                    "temperature": 0.1,  # 低温度以获得更确定性的回答
                    "num_predict": 2048  # 足够的token来生成回答
                }
            }

            response = requests.post(self.api_url, json=payload)
            response.raise_for_status()
            result = response.json()

            # 解析模型输出
            model_output = result.get('response', '')
            structured_data = self._parse_model_output(model_output)

            logger.info("LLM成功提取结构化信息")
            return structured_data

        except requests.exceptions.RequestException as e:
            logger.error(f"调用Ollama API失败: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"处理LLM响应时出错: {str(e)}")
            raise

    def _build_prompt(self, text: str, fields: List[str]) -> str:
        """
        Build the prompt for the LLM

        Args:
            text: Web page text content
            fields: List of fields to extract

        Returns:
            Constructed prompt
        """
        # Categorize different types of fields
        category_fields = ["Master Student", "Doctoral Student", "PostDoc", "Research Assistant",
                           "Competition", "Summer School", "Conference", "Workshop",
                           "Physical_Geo", "Human_Geo", "Urban", "GIS", "RS", "GNSS"]

        date_fields = ["Deadline"]

        number_fields = ["Number_Places"]

        # Columns not to be filled
        exclude_fields = ["Notes", "Source", "Verifier", "Error"]

        # All other fields are treated as text fields
        text_fields = [f for f in fields if f not in category_fields and
                       f not in date_fields and f not in number_fields and
                       f not in exclude_fields]

        # Create field list strings (for prompt)
        # Filter out exclude_fields from fields
        fields_to_extract = [f for f in fields if f not in exclude_fields]
        all_fields_str = ", ".join(fields_to_extract)
        category_fields_str = ", ".join(category_fields)
        date_fields_str = ", ".join(date_fields)
        number_fields_str = ", ".join(number_fields)
        text_fields_str = ", ".join(text_fields)

        prompt = f"""# Academic Opportunity Information Extraction Task
I need you to accurately extract structured information about academic opportunities from the following web page text. Please strictly follow the rules for each field type.

## Fields to Extract
{all_fields_str}

## Field Types and Processing Rules
These fields are divided into four different types, each with specific processing rules:

### 1. Category Fields (Fill with "1" or leave empty)
Applicable fields: {category_fields_str}

Processing rules:
- If the web page clearly indicates that the option applies or is available, fill with "1"
- If the web page clearly indicates that the option does not apply, or is not mentioned, leave empty (empty string "")
- These fields can only be filled with "1" or left empty, do not use "Yes", "No" or any other values
- If uncertain, default to empty
- Example: If the web page mentions that the opportunity is suitable for doctoral students, fill "1" in the "Doctoral Student" field
- Research domain categories (such as Physical_Geo, GIS, etc.) follow the same rules, fill "1" if the web page mentions the relevant field

### 2. Date Fields
Applicable fields: {date_fields_str}

Processing rules:
- Extract complete date information (year, month, day)
- Extract dates as they appear in the original text, without modifying the format
- For Deadline, look for application deadline, submission deadline, etc.
- If the date cannot be determined, leave empty

### 3. Quantity Fields (Fill with specific numbers)
Applicable fields: {number_fields_str}

Processing rules:
- Extract the specific number of positions or vacancies
- Must be Arabic numerals (e.g., "3", "10", etc.), not English words (e.g., not "one", "two", etc.)
- If the text contains numbers in word form (e.g., "One position"), convert them to Arabic numerals (e.g., "1")
- If multiple different positions are mentioned, add their quantities to get a total
- If the text clearly mentions a quantity but not as a specific number (e.g., "several", "multiple"), try to estimate and convert to a number
- If the quantity cannot be determined at all, leave empty

### 4. Text Content Fields
Applicable fields: {text_fields_str}

Processing rules:
- Extract the relevant specific text content, maintaining the accuracy of the original text
- Do not simplify to "1" or other symbols, must be actual text content
- Direction: Extract research direction or project description
- University_EN: English university name
- University_CN: Chinese university name (if any)
- Country_CN: Chinese country name
- Contact_Name: Contact person's name
- Contact_Email: Contact person's email
- WX_Label1-5: Possible tags or keywords, one keyword per tag, up to 5 tags
- If the relevant information cannot be found, leave empty

## Important Notes
1. Please thoroughly analyze the text, especially PDF content, to ensure no relevant information is missed
2. For numerical representations, always use Arabic numerals rather than English words, e.g., use "1" not "One"

## Output Format Requirements
- Use standard JSON format
- Field names must exactly match the provided field names (case-sensitive)
- Do not add any comments, explanations, or prefixes
- Ensure the output can be parsed as valid JSON

## Web Page Text Content
{text[:10000]}

## Output Example
```json
{{
  "Deadline": "2025-09-30",
  "Number_Places": "1",
  "Direction": "Climate change and ecosystem research",
  "University_EN": "University of Cambridge",
  "University_CN": "剑桥大学",
  "Country_CN": "英国",
  "Master Student": "1",
  "Doctoral Student": "1",
  "PostDoc": "",
  "Research Assistant": "",
  "Competition": "",
  "Summer School": "",
  "Conference": "",
  "Workshop": "",
  "Physical_Geo": "1",
  "Human_Geo": "",
  "Urban": "",
  "GIS": "1",
  "RS": "",
  "GNSS": "",
  "Contact_Name": "John Smith",
  "Contact_Email": "js123@cam.ac.uk",
  "WX_Label1": "Ecology",
  "WX_Label2": "Climate Change",
  "WX_Label3": "Geographic Information Systems",
  "WX_Label4": "",
  "WX_Label5": ""
}}
        """
        return prompt

    def _parse_model_output(self, output: str) -> Dict[str, str]:
        """
        解析模型输出为结构化数据

        Args:
            output: 模型输出的文本

        Returns:
            解析后的结构化数据字典
        """
        # 提取JSON部分
        try:
            # 尝试找到JSON部分（有时模型会在JSON前后添加额外文本）
            json_start = output.find('{')
            json_end = output.rfind('}') + 1

            if json_start >= 0 and json_end > json_start:
                json_str = output[json_start:json_end]
                data = json.loads(json_str)
                return data
            else:
                # 如果找不到JSON格式，尝试直接解析整个输出
                return json.loads(output)

        except json.JSONDecodeError:
            logger.error(f"无法解析LLM输出为JSON: {output}")
            # 返回空字典，避免程序崩溃
            return {}


class ExtractionSystem:
    """信息提取系统：整合各个模块，协调工作流程"""

    def __init__(self, excel_path: str, sheet_name: str = "Unfilled", model_name: str = "llama3"):
        """
        初始化信息提取系统

        Args:
            excel_path: Excel文件路径
            sheet_name: 工作表名称
            model_name: Ollama模型名称
        """
        self.excel_processor = ExcelProcessor(excel_path, sheet_name)
        self.web_extractor = WebPageExtractor()
        self.llm_processor = LlmProcessor(model_name=model_name)

    def run(self, url_column: str = "Source", batch_size: int = 10, fields: Optional[List[str]] = None) -> None:
        """
        运行信息提取系统

        Args:
            url_column: 包含URL的列名
            batch_size: 批处理大小
            fields: 需要提取的字段列表，如果为None则使用Excel中除了特殊列外的所有列
        """
        try:
            # 加载并过滤Excel数据
            self.excel_processor.load_data()
            filtered_df = self.excel_processor.filter_rows()

            if filtered_df.empty:
                logger.info("没有需要处理的行")
                return

            # 如果未提供字段列表，则使用Excel中除了特殊列外的所有列
            if fields is None:
                # 排除特殊列和URL列
                special_columns = ['Notes', 'Verifier', 'Error', url_column, '处理时间']
                fields = [col for col in filtered_df.columns if col not in special_columns]

            # 处理每一行
            for i, (idx, row) in enumerate(filtered_df.iterrows()):
                try:
                    # 检查Notes列中是否有URL，优先使用Notes列中的URL
                    notes_value = row['Notes'] if pd.notna(row['Notes']) else ""
                    if isinstance(notes_value, str) and self.excel_processor._is_url(notes_value):
                        url = notes_value.strip()
                        logger.info(f"使用Notes列中的URL: {url}")
                    else:
                        url = row[url_column]
                        if not url or pd.isna(url):
                            logger.warning(f"行 {idx} 的URL为空，跳过")
                            continue
                        logger.info(f"使用{url_column}列中的URL: {url}")

                    logger.info(f"处理行 {idx}, URL: {url}")

                    # 提取网页内容
                    web_content = self.web_extractor.extract_content(url)

                    # 使用LLM提取结构化信息
                    structured_info = self.llm_processor.extract_structured_info(web_content, fields)

                    # 将结果写回Excel
                    self.excel_processor.write_results(idx, structured_info)

                    # 每处理batch_size行后，休息一下
                    if (i + 1) % batch_size == 0:
                        logger.info(f"已处理 {i + 1} 行，休息5秒")
                        time.sleep(5)

                except Exception as e:
                    error_msg = f"处理失败: {str(e)}"
                    logger.error(f"处理行 {idx} 时出错: {str(e)}")
                    self.excel_processor.write_results(idx, {}, has_error=True, error_msg=error_msg)

            logger.info("所有行处理完成")
        except Exception as e:
            logger.error(f"运行过程中发生错误: {str(e)}")


def main():
    """主函数"""
    import argparse
    parser = argparse.ArgumentParser(description='LLM自动化信息提取系统')
    parser.add_argument('--excel', default='text_info.xlsx', help='Excel文件路径')
    parser.add_argument('--sheet', default='Unfilled', help='工作表名称')
    parser.add_argument('--model', default='qwen2.5vl:7b', help='Ollama模型名称')
    parser.add_argument('--url_column', default='Source', help='包含URL的列名')
    parser.add_argument('--batch_size', type=int, default=10, help='批处理大小')
    parser.add_argument('--fields', nargs='*', help='需要提取的字段列表')

    args = parser.parse_args()

    logger.info("启动LLM自动化信息提取系统")

    try:
        system = ExtractionSystem(
            excel_path=args.excel,
            sheet_name=args.sheet,
            model_name=args.model
        )

        system.run(
            url_column=args.url_column,
            batch_size=args.batch_size,
            fields=args.fields
        )

        logger.info("信息提取任务完成")

    except Exception as e:
        logger.error(f"系统运行出错: {str(e)}")


# 添加入口点
if __name__ == "__main__":
    main()
